import os.path

from django.db import models
from django.db.models.signals import post_save, post_delete
from django.dispatch import receiver
from django.contrib.auth.models import User
import random
import string
from simple_history.models import HistoricalRecords
import shlex

# Create your models here.
class Organisation(models.Model):
    name = models.CharField(max_length=100)
    kurzname = models.CharField(max_length=50, blank=True, null=True)
    email = models.EmailField()
    adress = models.TextField(null=True, blank=True)
    telefon = models.CharField(max_length=20, null=True, blank=True)
    website = models.URLField(null=True, blank=True)
    logo = models.ImageField(upload_to='logos/')
    farbe = models.CharField(max_length=7, default='#007bff')
    text_color_on_org_color = models.CharField(max_length=7, default='#000000')

    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Organisation'
        verbose_name_plural = 'Organisationen'

    def __str__(self):
        return self.name
    
@receiver(post_save, sender=Organisation)
def create_folder(sender, instance, created, **kwargs):
    if created:
        from Global.models import CustomUser, PersonCluster
        from ORG.tasks import send_register_email_task

        # Sanitize organization name for folder creation
        safe_org_name = instance.name.replace('/', '').replace('\\', '').replace('..', '')
        path = os.path.join(safe_org_name)
        os.makedirs(os.path.join('dokument', safe_org_name), exist_ok=True)

        if instance.kurzname:
            user_name = instance.kurzname.lower().replace(' ', '_')
        else:
            user_name = instance.name.lower().replace(' ', '_')

        user = User.objects.create(username=user_name, email=instance.email)
        user.is_superuser = True
        user.is_staff = True
        user.save()
        person_cluster = PersonCluster.objects.create(name=instance.name, org=instance, view='O')

        import random
        import string
        
        # Generate random string with letters and digits
        random_password = ''.join(random.choices(string.ascii_letters + string.digits, k=20))
        einmalpasswort = random.randint(10000000, 99999999)
        user.set_password(random_password)
        user.save()

        customuser = CustomUser.objects.create(user=user, org=instance, einmalpasswort=einmalpasswort, person_cluster=person_cluster)

        send_register_email_task.s(customuser.id).apply_async(countdown=10)


class MailBenachrichtigungen(models.Model):
    org = models.ForeignKey(Organisation, on_delete=models.CASCADE)
    betreff = models.CharField(max_length=100)
    text = models.TextField()

    class Meta:
        verbose_name = 'Mail Benachrichtigung'
        verbose_name_plural = 'Mail Benachrichtigungen'

    def __str__(self):
        return f'{self.organisation} - {self.betreff}'


class JahrgangTyp(models.Model):
    org = models.ForeignKey(Organisation, on_delete=models.CASCADE)
    name = models.CharField(max_length=50, verbose_name='Jahrgangstyp')

    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Jahrgangstyp'
        verbose_name_plural = 'Jahrgangstypen'

    def __str__(self):
        return self.name
    

class Ordner(models.Model):
    org = models.ForeignKey(Organisation, on_delete=models.CASCADE)
    ordner_name = models.CharField(max_length=100)
    typ = models.ForeignKey(JahrgangTyp, on_delete=models.SET_NULL, null=True, blank=True, verbose_name='Typ')
    color = models.ForeignKey('DokumentColor', on_delete=models.SET_NULL, null=True, blank=True, verbose_name='Farbe')

    history = HistoricalRecords()

    def __str__(self):
        return self.ordner_name


@receiver(post_save, sender=Ordner)
def create_folder(sender, instance, **kwargs):
    # Sanitize folder and org names
    safe_ordner_name = instance.ordner_name.replace('/', '').replace('\\', '').replace('..', '')
    safe_org_name = instance.org.name.replace('/', '').replace('\\', '').replace('..', '')
    
    path = os.path.join(safe_ordner_name)
    os.makedirs(os.path.join('dokument', safe_org_name, path), exist_ok=True)


@receiver(post_delete, sender=Ordner)
def remove_folder(sender, instance, **kwargs):
    # Sanitize folder and org names
    safe_ordner_name = instance.ordner_name.replace('/', '').replace('\\', '').replace('..', '')
    safe_org_name = instance.org.name.replace('/', '').replace('\\', '').replace('..', '')
    
    path = os.path.join(safe_ordner_name)
    path = os.path.join('dokument', safe_org_name, path)
    if os.path.isdir(path):
        os.rmdir(path)


def upload_to_folder(instance, filename):
    import os
    # Sanitize filename to prevent path traversal
    filename = os.path.basename(filename)
    # Remove any remaining path separators
    filename = filename.replace('/', '').replace('\\', '')
    
    order = instance.ordner
    # Sanitize folder name as well
    safe_ordner_name = order.ordner_name.replace('/', '').replace('\\', '').replace('..', '')
    safe_org_name = instance.org.name.replace('/', '').replace('\\', '').replace('..', '')
    
    path = os.path.join(safe_ordner_name, filename)
    return os.path.join('dokument', safe_org_name, path)


def upload_to_preview_image(instance, filename):
    import os
    # Sanitize filename
    filename = os.path.basename(filename)
    filename = filename.replace('/', '').replace('\\', '')
    
    filename = filename.split('/')[-1]
    filename = filename.split('.')[0]
    
    # Sanitize org name
    safe_org_name = instance.org.name.replace('/', '').replace('\\', '').replace('..', '')
    
    folder = os.path.join('dokument', safe_org_name, 'preview_image')
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, filename + '.jpg')


class Dokument(models.Model):
    org = models.ForeignKey(Organisation, on_delete=models.CASCADE)
    ordner = models.ForeignKey(Ordner, on_delete=models.CASCADE)
    dokument = models.FileField(upload_to=upload_to_folder, max_length=255, null=True, blank=True)
    link = models.URLField(null=True, blank=True)
    date_created = models.DateTimeField(auto_now_add=True)
    date_modified = models.DateTimeField(auto_now=True)
    titel = models.CharField(max_length=100, null=True, blank=True)
    beschreibung = models.TextField(null=True, blank=True)
    fw_darf_bearbeiten = models.BooleanField(default=True)
    preview_image = models.ImageField(upload_to=upload_to_preview_image, null=True, blank=True)

    history = HistoricalRecords()

    def __str__(self):
        return self.titel or self.dokument.name or self.link

    def get_document_type(self):
        if self.dokument:
            import mimetypes
            file_path = self.dokument.path
            # Guess the MIME type based on file extension
            mime_type, _ = mimetypes.guess_type(file_path)

            return mime_type or 'unknown'
        else:
            return 'unknown'
        
    def get_document_suffix(self):
        if self.dokument:
            return self.dokument.name.split('.')[-1]
        else:
            return 'unknown'
        
    def get_preview_image(self):
        if self.preview_image and os.path.exists(self.preview_image.path):
            return self.preview_image.path
        else:
            return self.get_preview_converted()
            
    
    def get_preview_converted(self):
        import subprocess
        import hashlib

        def pdf_to_image(doc_path, img_path):
                from pdf2image import convert_from_path
                image = convert_from_path(doc_path, first_page=1, last_page=1)[0]
                image.save(img_path)
                if os.path.exists(img_path):
                    return img_path
                else:
                    return None
            
        def excel_to_image(excel_path, img_path):
            from openpyxl import load_workbook
            from PIL import Image, ImageDraw, ImageFont
            import numpy as np

            # Load the workbook
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            # Define image parameters
            cell_width = 100
            cell_height = 30
            max_rows = 10
            max_cols = 6
            padding = 10

            # Calculate actual dimensions needed
            used_rows = min(ws.max_row, max_rows)
            used_cols = min(ws.max_column, max_cols)

            # Create image with white background
            img_width = (cell_width * used_cols) + padding * 2
            img_height = (cell_height * (used_rows + 1)) + padding * 2  # +1 for header
            img = Image.new('RGB', (img_width, img_height), 'white')
            draw = ImageDraw.Draw(img)

            try:
                # Try to load Arial font, fall back to default if not available
                font = ImageFont.truetype('Arial', 12)
            except:
                font = ImageFont.load_default()

            # Draw grid and cell contents
            for row in range(used_rows):
                for col in range(used_cols):
                    # Calculate cell position
                    x1 = col * cell_width + padding
                    y1 = row * cell_height + padding
                    x2 = x1 + cell_width
                    y2 = y1 + cell_height

                    # Draw cell border
                    draw.rectangle([x1, y1, x2, y2], outline='gray')

                    # Get cell value
                    cell = ws.cell(row=row+1, column=col+1)
                    value = str(cell.value if cell.value is not None else '')
                    if len(value) > 15:
                        value = value[:12] + '...'

                    # Draw text
                    draw.text((x1 + 5, y1 + 5), value, fill='black', font=font)

            # Save the image
            img.save(img_path)
            return img_path
            
        def get_hashed_filename(filename):
            """Create a shorter, hashed filename while preserving extension"""
            name, ext = os.path.splitext(filename)
            hash_object = hashlib.md5(name.encode())
            hashed_name = hash_object.hexdigest()[:8]  # Use first 8 chars of hash
            return f"{hashed_name}{ext}"
        
        if not self.dokument:
            return None
            
        mimetype = self.get_document_type()
        
        # Create hashed filename for preview image
        hashed_name = get_hashed_filename(self.dokument.name)
        preview_image_path = upload_to_preview_image(self, hashed_name + '.jpg')

        if mimetype and mimetype.startswith('image'):
            return self.dokument.path

        if mimetype and mimetype == 'application/pdf':
            return pdf_to_image(self.dokument.path, preview_image_path)

        elif self.dokument.name.endswith('.docx') or self.dokument.name.endswith('.doc') or self.dokument.name.endswith('.odt'):
            # When shell=False (default), subprocess.run safely handles arguments
            command = ["abiword", "--to=pdf", self.dokument.path]
            try:
                subprocess.run(command, check=True, capture_output=True, timeout=30)
                doc_path = str(self.dokument.path)  # Create string copy
                if self.dokument.name.endswith('.docx'):
                    doc_path = doc_path.replace('.docx', '.pdf')
                elif self.dokument.name.endswith('.odt'):
                    doc_path = doc_path.replace('.odt', '.pdf')
                elif self.dokument.name.endswith('.doc'):
                    doc_path = doc_path.replace('.doc', '.pdf')
                pdf_to_image(doc_path, preview_image_path)
            except Exception as e:
                print(e)
                return None

        elif self.dokument.name.endswith('.xlsx') or self.dokument.name.endswith('.xls'):
            try:
                return excel_to_image(self.dokument.path, preview_image_path)
            except Exception as e:
                print(f"Error creating Excel preview: {e}")
                return None
            
        if os.path.exists(preview_image_path):
            if not self.preview_image:
                self.preview_image = preview_image_path
                self.save()
            return preview_image_path
        else:
            return None
        
@receiver(post_save, sender=Dokument)
def create_preview_image(sender, instance, **kwargs):
    # Skip if we're already processing the preview image
    if hasattr(instance, '_creating_preview'):
        return
    
    img_path = instance.get_preview_converted()
    if img_path:
        try:
            # Set flag to prevent recursive save
            instance._creating_preview = True
            instance.preview_image = img_path
            instance.save()
        finally:
            # Always remove the flag, even if an error occurs
            delattr(instance, '_creating_preview')

@receiver(post_delete, sender=Dokument)
def remove_file(sender, instance, **kwargs):
    if instance.dokument and os.path.isfile(instance.dokument.path):
        os.remove(instance.dokument.path)

    if instance.preview_image and os.path.isfile(instance.preview_image.path):
        os.remove(instance.preview_image.path)

class DokumentColor(models.Model):
    name = models.CharField(max_length=50, verbose_name='Farbname')
    color = models.CharField(max_length=7, verbose_name='Farbcodes')

    history = HistoricalRecords()

    def __str__(self):
        return self.name

class Referenten(models.Model):
    org = models.ForeignKey(Organisation, on_delete=models.CASCADE)
    user = models.OneToOneField(User, on_delete=models.CASCADE, blank=True, null=True)
    first_name = models.CharField(max_length=50, verbose_name='Vorname', null=True, blank=True)
    last_name = models.CharField(max_length=50, verbose_name='Nachname')
    email = models.EmailField(verbose_name='E-Mail')
    phone_work = models.CharField(max_length=20, verbose_name='Telefon Arbeit', null=True, blank=True)
    phone_mobil = models.CharField(max_length=20, verbose_name='Telefon Mobil', null=True, blank=True)

    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Teammitglied'
        verbose_name_plural = 'Team'

    def __str__(self):
        return f'{self.last_name}, {self.first_name}'

@receiver(post_save, sender=Referenten)
def create_user(sender, instance, **kwargs):
    if not instance.user:
        from Global.models import CustomUser

        default_username = instance.last_name.lower().replace(' ', '_')
        user = User.objects.create(username=default_username, email=instance.email)
        random_password = ''.join(random.choices(string.ascii_letters + string.digits, k=20))
        user.set_password(random_password)
        user.save()
    
        einmalpasswort = random.randint(10000000, 99999999)
        customuser = CustomUser.objects.create(user=user, org=instance.org, role='T', einmalpasswort=einmalpasswort)
        
        instance.user = user
        instance.save()