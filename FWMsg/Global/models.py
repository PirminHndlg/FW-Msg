from django.db import models
from django.contrib.auth.models import User
from django.dispatch import receiver
from django.core import validators
import random
import string
from Global.tasks import send_new_post_email_task
from simple_history.models import HistoricalRecords
import os.path
from django.db.models.signals import post_save, post_delete, pre_delete
from FWMsg.middleware import get_current_request
import os
from ORG.models import Organisation
from datetime import datetime, timedelta
from django.core import signing
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
import logging

logger = logging.getLogger(__name__)

from PIL import Image, ImageOps  # Make sure this is from PIL, not Django models
from django.core.files.base import ContentFile
import io
from django.urls import reverse


class OrgManager(models.Manager):
    def get_queryset(self):
        request = get_current_request()
        if request and hasattr(request, 'user') and hasattr(request.user, 'org') and not request.user.is_superuser:
            return super().get_queryset().filter(org=request.user.org)
        return super().get_queryset()



class OrgModel(models.Model):
    org = models.ForeignKey(Organisation, on_delete=models.CASCADE, verbose_name='Organisation')
    objects = OrgManager()

    class Meta:
        abstract = True


class PersonCluster(OrgModel):
    view_choices = [
        ('A', 'Admin'),
        ('O', 'Organisation'),
        ('F', 'Freiwillige:r'),
        ('E', 'Ehemalige:r'),
        ('T', 'Team'),
        ('B', 'Bewerber:in')
    ]

    name = models.CharField(max_length=50, verbose_name='Benutzergruppe', help_text='Name der Benutzergruppe, z.B. "Freiwillige 2023" oder "Teamleitung"')

    aufgaben = models.BooleanField(default=False, verbose_name='Aufgaben anzeigen', help_text='Aktivieren, um Aufgaben für diese Gruppe anzuzeigen')
    calendar = models.BooleanField(default=False, verbose_name='Kalender anzeigen', help_text='Aktivieren, um den Kalender für diese Gruppe anzuzeigen')
    dokumente = models.BooleanField(default=False, verbose_name='Dokumente anzeigen', help_text='Aktivieren, um Dokumente für diese Gruppe anzuzeigen')
    ampel = models.BooleanField(default=False, verbose_name='Ampel anzeigen', help_text='Aktivieren, um die Ampelfunktion für diese Gruppe anzuzeigen')
    notfallkontakt = models.BooleanField(default=False, verbose_name='Notfallkontakt anzeigen', help_text='Aktivieren, um Notfallkontakte für diese Gruppe anzuzeigen')
    bilder = models.BooleanField(default=False, verbose_name='Bilder anzeigen', help_text='Aktivieren, um Bilder für diese Gruppe anzuzeigen')
    posts = models.BooleanField(default=False, verbose_name='Posts anzeigen', help_text='Aktivieren, um Posts für diese Gruppe anzuzeigen')

    view = models.CharField(max_length=1, choices=view_choices, default='F', verbose_name='Standardansicht', help_text='Bestimmt die Standardansicht für Mitglieder dieser Gruppe')

    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Benutzergruppe'
        verbose_name_plural = 'Benutzergruppen'
        
    def __str__(self):
        return self.name
    
    def get_users(self):
        return User.objects.filter(customuser__person_cluster=self)
    

class CustomUser(OrgModel):
    user = models.OneToOneField(User, on_delete=models.CASCADE, verbose_name='Benutzer:in')
    person_cluster = models.ForeignKey(PersonCluster, on_delete=models.CASCADE, null=True, blank=True, verbose_name='Benutzergruppe')
    profil_picture = models.ImageField(upload_to='profil_picture/', blank=True, null=True, verbose_name='Profilbild')
    geburtsdatum = models.DateField(blank=True, null=True, verbose_name='Geburtsdatum')

    mail_notifications = models.BooleanField(default=True, verbose_name='Mail-Benachrichtigungen')
    mail_notifications_unsubscribe_auth_key = models.CharField(max_length=255, blank=True, null=True, verbose_name='Mail-Benachrichtigung Abmelde-Key')

    einmalpasswort = models.CharField(max_length=20, blank=True, null=True, verbose_name='Einmalpasswort', help_text='Wird automatisch erzeugt, wenn leer')
    token = models.CharField(max_length=512, blank=True, null=True, verbose_name='Token', help_text='Wird automatisch erzeugt, wenn leer')
    calendar_token = models.CharField(max_length=512, blank=True, null=True, verbose_name='Kalender-Token', help_text='Wird automatisch erzeugt, wenn leer')
    
    # Online status tracking
    last_seen = models.DateTimeField(blank=True, null=True, verbose_name='Zuletzt online')
    is_online = models.BooleanField(default=False, verbose_name='Ist online')

    history = HistoricalRecords()

    def delete(self, *args, **kwargs):
        # Delete the associated User first
        if self.user:
            self.user.delete()
        # Then delete the CustomUser
        super().delete(*args, **kwargs)

    def send_registration_email(self):
        if not self.einmalpasswort:
            self.einmalpasswort = random.randint(100000, 999999)
            self.save()

        if self.person_cluster.view == 'F':
            from FW.tasks import send_register_email_task
            send_register_email_task.s(self.id).apply_async(countdown=2)
        else:
            from ORG.tasks import send_register_email_task
            send_register_email_task.s(self.id).apply_async(countdown=2)
        
    def get_random_hash(self, length=16):
        import secrets
        import hashlib
        import time
        
        # Combine user ID with current timestamp and a random token
        unique_value = f"{self.user.id}_{time.time()}_{secrets.token_hex(length)}"
        
        # Create a SHA-256 hash of this value
        hash_obj = hashlib.sha512(unique_value.encode())
        return hash_obj.hexdigest()

    def get_unsubscribe_url(self):
        try:
            if not self.mail_notifications_unsubscribe_auth_key:
                self.mail_notifications_unsubscribe_auth_key = self.get_random_hash()
                self.save()
            
            # Use Django's reverse function instead of hardcoding the path
            relative_url = reverse('unsubscribe_mail_notifications', kwargs={
                'user_id': self.user.id,
                'auth_key': self.mail_notifications_unsubscribe_auth_key
            })
            
            # Add the base domain to the relative URL
            base_url = "https://volunteer.solutions"
            return base_url + relative_url
        except Exception as e:
            print(e)
            # Fallback to profile page if there's an error
            return "https://volunteer.solutions/profil"

    def create_small_image(self):
        if self.profil_picture:
            self.profil_picture = calculate_small_image(self.profil_picture, size=(500, 500))
            self.save()
            
    def create_token(self):
        self.token = self.get_random_hash(128)
        self.save()
        
    def create_calendar_token(self):
        self.calendar_token = self.get_random_hash(128)
        self.save()

    def ensure_token(self):
        """Ensure the user has a valid token for calendar subscription."""
        if not self.token:
            self.create_token()
        return self.token

    def ensure_calendar_token(self):
        """Ensure the user has a valid token for calendar subscription."""
        if not self.calendar_token:
            self.create_calendar_token()
        return self.calendar_token
    
    def update_last_seen(self):
        """Update the last seen timestamp and online status."""
        from django.utils import timezone
        self.last_seen = timezone.now()
        self.is_online = True
        self.save(update_fields=['last_seen', 'is_online'])
    
    def is_currently_online(self):
        """Check if user is currently online (active within last 5 minutes)."""
        if not self.last_seen:
            return False
        from django.utils import timezone
        from datetime import timedelta
        return timezone.now() - self.last_seen < timedelta(minutes=5)
    
    def get_online_status_display(self):
        """Get a human-readable online status."""
        if self.is_currently_online():
            return "Online"
        elif self.last_seen:
            from django.utils import timezone
            from datetime import timedelta
            time_diff = timezone.now() - self.last_seen
            
            if time_diff.days > 0:
                return f"Vor {time_diff.days} Tag{'en' if time_diff.days > 1 else ''}"
            elif time_diff.seconds > 3600:
                hours = time_diff.seconds // 3600
                return f"Vor {hours} Stunde{'n' if hours > 1 else ''}"
            elif time_diff.seconds > 60:
                minutes = time_diff.seconds // 60
                return f"Vor {minutes} Minute{'n' if minutes > 1 else ''}"
            else:
                return "Gerade eben"
        else:
            return "Nie online gewesen"

    def __str__(self):
        return self.user.username
    
    class Meta:
        verbose_name = 'Benutzer:in'
        verbose_name_plural = 'Benutzer:innen'

    
@receiver(post_save, sender=CustomUser)
def post_save_handler(sender, instance, created, **kwargs):

    if instance.profil_picture and not hasattr(instance, '_processing_profil_picture'):
        instance._processing_profil_picture = True
        instance.create_small_image()
        delattr(instance, '_processing_profil_picture')

# Add property to User model to access org
User.add_to_class('org', property(lambda self: self.customuser.org if hasattr(self, 'customuser') else None))
User.add_to_class('view', property(lambda self: self.customuser.person_cluster.view if hasattr(self, 'customuser') and self.customuser.person_cluster else None))
User.add_to_class('role', property(lambda self: self.customuser.person_cluster.view if hasattr(self, 'customuser') and self.customuser.person_cluster else None))
User.add_to_class('person_cluster', property(lambda self: self.customuser.person_cluster if hasattr(self, 'customuser') and self.customuser.person_cluster else None))

def user_str_method(self):
    if hasattr(self, 'first_name') and hasattr(self, 'last_name') and self.first_name and self.last_name:
        return f"{self.first_name} {self.last_name}"
    return self.username

User.__str__ = user_str_method

class Feedback(models.Model):
    text = models.TextField(verbose_name=_('Feedback-Text'))
    user = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name=_('Benutzer:in'), null=True, blank=True)
    anonymous = models.BooleanField(default=False, verbose_name=_('Anonym einreichen'))

    def __str__(self):
        return f'Feedback von {self.user.username}' if not self.anonymous and self.user else 'Anonymes Feedback'

@receiver(post_save, sender=Feedback)
def post_save_handler(sender, instance, created, **kwargs):
    if instance.anonymous and instance.user:
        instance.user = None
        instance.save()
    
    if created:
        from ORG.tasks import send_feedback_email_task
        send_feedback_email_task.s(instance.id).apply_async(countdown=10)

class KalenderEvent(OrgModel):
    user = models.ManyToManyField(User, verbose_name='Teilnehmer:innen', help_text='Personen, die an diesem Termin teilnehmen', related_name='calendar_events')
    title = models.CharField(max_length=255, verbose_name='Titel', help_text='Titel des Termins')
    start = models.DateTimeField(verbose_name='Beginnt am', help_text='Startdatum und -uhrzeit des Termins')
    end = models.DateTimeField(verbose_name='Endet am', help_text='Enddatum und -uhrzeit des Termins')
    description = models.TextField(verbose_name='Beschreibung', null=True, blank=True, help_text='Ausführliche Beschreibung des Termins (optional)')
    mail_reminder_sent_to = models.ManyToManyField(User, blank=True, verbose_name='Mail-Erinnerung gesendet an', help_text='Personen, an die eine Mail-Erinnerung gesendet wurde', related_name='calendar_reminders')

    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Kalender-Termin'
        verbose_name_plural = 'Kalender-Termine'

    def __str__(self):
        return self.title
    
    
class Ordner2(OrgModel):
    ordner_name = models.CharField(max_length=255, verbose_name='Ordnername', help_text='Name des Ordners')
    typ = models.ManyToManyField(PersonCluster, verbose_name='Sichtbar für', help_text='Benutzergruppen, die diesen Ordner sehen können')
    color = models.ForeignKey('DokumentColor2', on_delete=models.SET_NULL, null=True, blank=True, verbose_name='Farbkennzeichnung', help_text='Farbliche Markierung des Ordners')

    history = HistoricalRecords()

    def __str__(self):
        return self.ordner_name


@receiver(post_save, sender=Ordner2)
def create_folder(sender, instance, **kwargs):
    # Sanitize folder and org names
    safe_ordner_name = instance.ordner_name.replace('/', '').replace('\\', '').replace('..', '')
    safe_org_name = instance.org.name.replace('/', '').replace('\\', '').replace('..', '')
    
    path = os.path.join(safe_ordner_name)
    os.makedirs(os.path.join('dokument', safe_org_name, path), exist_ok=True)


@receiver(post_delete, sender=Ordner2)
def remove_folder(sender, instance, **kwargs):
    # Sanitize folder and org names
    safe_ordner_name = instance.ordner_name.replace('/', '').replace('\\', '').replace('..', '')
    safe_org_name = instance.org.name.replace('/', '').replace('\\', '').replace('..', '')
    
    path = os.path.join(safe_ordner_name)
    path = os.path.join('dokument', safe_org_name, path)
    if os.path.isdir(path):
        os.rmdir(path)


def upload_to_folder(instance, filename):
    import os
    # Sanitize filename to prevent path traversal
    filename = os.path.basename(filename)
    # Remove any remaining path separators
    filename = filename.replace('/', '').replace('\\', '')
    
    order = instance.ordner
    # Sanitize folder name as well
    safe_ordner_name = order.ordner_name.replace('/', '').replace('\\', '').replace('..', '')
    safe_org_name = instance.org.name.replace('/', '').replace('\\', '').replace('..', '')
    
    path = os.path.join(safe_ordner_name, filename)
    return os.path.join('dokument', safe_org_name, path)


def upload_to_preview_image(instance, filename):
    import os
    # Sanitize filename
    filename = os.path.basename(filename)
    filename = filename.replace('/', '').replace('\\', '')
    
    filename = filename.split('/')[-1]
    filename = filename.split('.')[0]
    
    # Sanitize org name
    safe_org_name = instance.org.name.replace('/', '').replace('\\', '').replace('..', '')
    
    folder = os.path.join('dokument', safe_org_name, 'preview_image')
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, filename + '.jpg')


class Dokument2(models.Model):
    org = models.ForeignKey(Organisation, on_delete=models.CASCADE)
    ordner = models.ForeignKey(Ordner2, on_delete=models.CASCADE)
    dokument = models.FileField(upload_to=upload_to_folder, max_length=255, null=True, blank=True)
    link = models.URLField(null=True, blank=True, verbose_name='Link', help_text='Link zu einer externen Datei')
    date_created = models.DateTimeField(auto_now_add=True, verbose_name='Erstellt am', help_text='Datum der Erstellung der Datei')
    date_modified = models.DateTimeField(auto_now=True, verbose_name='Zuletzt geändert am', help_text='Datum der letzten Änderung der Datei')
    titel = models.CharField(max_length=255, null=True, blank=True, verbose_name='Titel', help_text='Titel der Datei')
    beschreibung = models.TextField(null=True, blank=True, verbose_name='Beschreibung', help_text='Beschreibung der Datei')
    darf_bearbeiten = models.ManyToManyField(PersonCluster, verbose_name='Darf bearbeiten', help_text='Benutzergruppen, die diese Datei bearbeiten können')
    preview_image = models.ImageField(upload_to=upload_to_preview_image, null=True, blank=True)

    history = HistoricalRecords()

    def __str__(self):
        return self.titel or self.dokument.name or self.link

    def get_document_type(self):
        if self.dokument:
            import mimetypes
            file_path = self.dokument.path
            # Guess the MIME type based on file extension
            mime_type, _ = mimetypes.guess_type(file_path)

            return mime_type or 'unknown'
        else:
            return 'unknown'
        
    def get_document_suffix(self):
        if self.dokument:
            return self.dokument.name.split('.')[-1]
        else:
            return 'unknown'
        
    def get_preview_image(self):
        if self.preview_image and os.path.exists(self.preview_image.path):
            return self.preview_image.path
        else:
            return self.get_preview_converted()
            
    
    def get_preview_converted(self):
        import subprocess
        import hashlib

        def pdf_to_image(doc_path, img_path):
                from pdf2image import convert_from_path
                image = convert_from_path(doc_path, first_page=1, last_page=1)[0]
                image.save(img_path)
                if os.path.exists(img_path):
                    return img_path
                else:
                    return None
            
        def excel_to_image(excel_path, img_path):
            from openpyxl import load_workbook
            from PIL import Image, ImageDraw, ImageFont
            import numpy as np

            # Load the workbook
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            # Define image parameters
            cell_width = 100
            cell_height = 30
            max_rows = 10
            max_cols = 6
            padding = 10

            # Calculate actual dimensions needed
            used_rows = min(ws.max_row, max_rows)
            used_cols = min(ws.max_column, max_cols)

            # Create image with white background
            img_width = (cell_width * used_cols) + padding * 2
            img_height = (cell_height * (used_rows + 1)) + padding * 2  # +1 for header
            img = Image.new('RGB', (img_width, img_height), 'white')
            draw = ImageDraw.Draw(img)

            try:
                # Try to load Arial font, fall back to default if not available
                font = ImageFont.truetype('Arial', 12)
            except:
                font = ImageFont.load_default()

            # Draw grid and cell contents
            for row in range(used_rows):
                for col in range(used_cols):
                    # Calculate cell position
                    x1 = col * cell_width + padding
                    y1 = row * cell_height + padding
                    x2 = x1 + cell_width
                    y2 = y1 + cell_height

                    # Draw cell border
                    draw.rectangle([x1, y1, x2, y2], outline='gray')

                    # Get cell value
                    cell = ws.cell(row=row+1, column=col+1)
                    value = str(cell.value if cell.value is not None else '')
                    if len(value) > 15:
                        value = value[:12] + '...'

                    # Draw text
                    draw.text((x1 + 5, y1 + 5), value, fill='black', font=font)

            # Save the image
            img.save(img_path)
            return img_path
            
        def get_hashed_filename(filename):
            """Create a shorter, hashed filename while preserving extension"""
            name, ext = os.path.splitext(filename)
            hash_object = hashlib.md5(name.encode())
            hashed_name = hash_object.hexdigest()[:8]  # Use first 8 chars of hash
            return f"{hashed_name}{ext}"
        
        if not self.dokument:
            return None
            
        mimetype = self.get_document_type()
        
        # Create hashed filename for preview image
        hashed_name = get_hashed_filename(self.dokument.name)
        preview_image_path = upload_to_preview_image(self, hashed_name + '.jpg')

        if mimetype and mimetype.startswith('image'):
            return self.dokument.path

        if mimetype and mimetype == 'application/pdf':
            return pdf_to_image(self.dokument.path, preview_image_path)

        elif self.dokument.name.endswith('.docx') or self.dokument.name.endswith('.doc') or self.dokument.name.endswith('.odt'):
            # When shell=False (default), subprocess.run safely handles arguments
            command = ["abiword", "--to=pdf", self.dokument.path]
            try:
                subprocess.run(command, check=True, capture_output=True, timeout=30)
                doc_path = str(self.dokument.path)  # Create string copy
                if self.dokument.name.endswith('.docx'):
                    doc_path = doc_path.replace('.docx', '.pdf')
                elif self.dokument.name.endswith('.odt'):
                    doc_path = doc_path.replace('.odt', '.pdf')
                elif self.dokument.name.endswith('.doc'):
                    doc_path = doc_path.replace('.doc', '.pdf')
                pdf_to_image(doc_path, preview_image_path)
            except Exception as e:
                print(e)
                return None

        elif self.dokument.name.endswith('.xlsx') or self.dokument.name.endswith('.xls'):
            try:
                return excel_to_image(self.dokument.path, preview_image_path)
            except Exception as e:
                print(f"Error creating Excel preview: {e}")
                return None
            
        if os.path.exists(preview_image_path):
            if not self.preview_image:
                self.preview_image = preview_image_path
                self.save()
            return preview_image_path
        else:
            return None
        
@receiver(post_save, sender=Dokument2)
def create_preview_image(sender, instance, **kwargs):
    # Skip if we're already processing the preview image
    if hasattr(instance, '_creating_preview'):
        return
    
    img_path = instance.get_preview_converted()
    if img_path:
        try:
            # Set flag to prevent recursive save
            instance._creating_preview = True
            instance.preview_image = img_path
            instance.save()
        except Exception as e:
            logger.error(f"Error creating preview image: {e}")
        finally:
            # Always remove the flag, even if an error occurs
            delattr(instance, '_creating_preview')

@receiver(post_delete, sender=Dokument2)
def remove_file(sender, instance, **kwargs):
    if instance.dokument and os.path.isfile(instance.dokument.path):
        os.remove(instance.dokument.path)

    if instance.preview_image and os.path.isfile(instance.preview_image.path):
        os.remove(instance.preview_image.path)

class DokumentColor2(models.Model):
    name = models.CharField(max_length=50, verbose_name='Farbname')
    color = models.CharField(max_length=7, verbose_name='Farbcodes')

    history = HistoricalRecords()

    def __str__(self):
        return self.name


def calculate_small_image(image, size=(750, 750), jpeg_quality=85, background_color=(255, 255, 255)):
    # Open the image
    img = Image.open(image)

    # Apply EXIF-based orientation correction (handles 1-8)
    try:
        img = ImageOps.exif_transpose(img)
    except Exception:
        pass

    # Ensure RGB for JPEG; handle transparency if present
    if img.mode not in ("RGB", "L"):
        try:
            if "A" in img.getbands():
                background = Image.new("RGB", img.size, background_color)
                background.paste(img, mask=img.split()[-1])
                img = background
            else:
                img = img.convert("RGB")
        except Exception:
            img = img.convert("RGB")

    # Create thumbnail with high-quality resampling
    try:
        img.thumbnail(size, resample=Image.LANCZOS)
    except Exception:
        img.thumbnail(size)

    # Encode as optimized/progressive JPEG
    img_io = io.BytesIO()
    img.save(img_io, format="JPEG", quality=jpeg_quality, optimize=True, progressive=True)

    # Ensure the returned filename has .jpg extension
    base = image.name.split('/')[-1] if hasattr(image, 'name') else 'image'
    base = os.path.splitext(base)[0] + '.jpg'
    return ContentFile(img_io.getvalue(), name=base)


class Einsatzland2(OrgModel):
    name = models.CharField(max_length=50, verbose_name='Einsatzland')
    code = models.CharField(max_length=2, verbose_name='Einsatzland-Code', help_text='Zweistelliger ISO-Code des Einsatzlandes (z.B. DE, CH, AT, ...)')

    notfallnummern = models.TextField(verbose_name='Notfallnummern', null=True, blank=True)
    arztpraxen = models.TextField(verbose_name='Arztpraxen', null=True, blank=True)
    apotheken = models.TextField(verbose_name='Apotheken', null=True, blank=True)
    informationen = models.TextField(verbose_name='Weitere Informationen', null=True, blank=True)

    history = HistoricalRecords()
    class Meta:
        verbose_name = 'Einsatzland'
        verbose_name_plural = 'Einsatzländer'
        ordering = ['name', 'code']

    def __str__(self):
        return self.name

class Einsatzstelle2(OrgModel):
    name = models.CharField(max_length=50, verbose_name='Einsatzstelle')
    land = models.ForeignKey(Einsatzland2, on_delete=models.CASCADE, verbose_name='Einsatzland', null=True, blank=True)

    partnerorganisation = models.TextField(verbose_name='Partnerorganisation', null=True, blank=True)
    arbeitsvorgesetzter = models.TextField(verbose_name='Arbeitsvorgesetzte:r', null=True, blank=True)
    mentor = models.TextField(verbose_name='Mentor:in', null=True, blank=True)
    botschaft = models.TextField(verbose_name='Botschaft', null=True, blank=True)
    konsulat = models.TextField(verbose_name='Konsulat', null=True, blank=True)
    informationen = models.TextField(verbose_name='Weitere Informationen', null=True, blank=True)
    
    max_freiwillige = models.IntegerField(blank=True, null=True, verbose_name='Maximale Anzahl Freiwillige')


    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Einsatzstelle'
        verbose_name_plural = 'Einsatzstellen'

    def __str__(self):
        return self.name
    

    def get_notiz_count(self):
        return EinsatzstelleNotiz.objects.filter(einsatzstelle=self).count()

class Attribute(OrgModel):
    TYPE_CHOICES = [
        ('T', 'Text'),
        ('L', 'Langer Text'),
        ('N', 'Zahl'),
        ('D', 'Datum'),
        ('B', 'Wahrheitswert'),
        ('E', 'E-Mail'),
        ('P', 'Telefon'),
        ('C', 'Auswahl')
    ]

    name = models.CharField(max_length=50, verbose_name='Bezeichnung', help_text='Name des benutzerdefinierten Datenfelds')
    type = models.CharField(max_length=1, choices=TYPE_CHOICES, verbose_name='Datentyp', default='T', help_text='Art der Daten, die in diesem Feld gespeichert werden')
    value_for_choices = models.CharField(null=True, blank=True, max_length=250, verbose_name='Auswahloptionen', help_text='Nur für Datentyp "Auswahl": Kommagetrennte Optionen, z.B. "Vegan, Vegetarisch, Konventionell"')
    person_cluster = models.ManyToManyField(PersonCluster, verbose_name='Für Benutzergruppen', help_text='Benutzergruppen, für die dieses Feld relevant ist')

    class Meta:
        verbose_name = 'Benutzerdatenfeld'
        verbose_name_plural = 'Benutzerdatenfelder'

    def __str__(self):
        return self.name


class UserAttribute(OrgModel):
    user = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name='Benutzer:in')
    attribute = models.ForeignKey(Attribute, on_delete=models.CASCADE, verbose_name='Benutzerdatenfeld')
    value = models.TextField(verbose_name='Wert', null=True, blank=True)

    class Meta:
        verbose_name = 'Freiwilliger Attribut'
        verbose_name_plural = 'Freiwilliger Attribute'

    def __str__(self):
        return self.user.first_name + ' ' + self.user.last_name + ' - ' + self.attribute.name


class Notfallkontakt2(OrgModel):
    first_name = models.CharField(max_length=50, verbose_name='Vorname', help_text='Vorname des Notfallkontakts')
    last_name = models.CharField(max_length=50, verbose_name='Nachname', help_text='Nachname des Notfallkontakts')
    phone_work = models.CharField(max_length=20, blank=True, null=True, verbose_name='Telefon (Arbeit)', help_text='Geschäftliche Telefonnummer')
    phone = models.CharField(max_length=20, blank=True, null=True, verbose_name='Telefon (Mobil)', help_text='Mobiltelefonnummer')
    email = models.EmailField(max_length=100, blank=True, null=True, verbose_name='E-Mail', help_text='E-Mail-Adresse des Notfallkontakts')
    user = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name='Benutzer:in', null=True, blank=True, help_text='Zugehörige:r Benutzer:in, für den dieser Notfallkontakt gilt')

    class Meta:
        verbose_name = 'Notfallkontakt'
        verbose_name_plural = 'Notfallkontakte'

    def __str__(self):
        return self.first_name + ' ' + self.last_name


class Ampel2(OrgModel):
    CHOICES = [
        ('G', 'Grün'),
        ('Y', 'Gelb'),
        ('R', 'Rot'),
    ]

    user = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name='Benutzer:in', help_text='Benutzer:in, für den dieser Status gilt')
    status = models.CharField(max_length=1, choices=CHOICES, verbose_name='Status', help_text='Aktueller Status (Grün = alles in Ordnung, Gelb = Aufmerksamkeit nötig, Rot = unmittelbare Hilfe erforderlich)')
    comment = models.TextField(blank=True, null=True, verbose_name='Kommentar', help_text='Kommentar zur Erläuterung des Status')
    date = models.DateTimeField(auto_now_add=True, verbose_name='Erstellt am')

    class Meta:
        verbose_name = 'Ampel'
        verbose_name_plural = 'Ampeln'

    def __str__(self):
        return self.user.first_name + ' ' + self.user.last_name + ' - ' + self.status


class AufgabenCluster(OrgModel):
    TYPE_CHOICES = [
        ('V', 'Vor Einsatz'),
        ('W', 'Während Einsatz'),
        ('N', 'Nach Einsatz'),
    ]

    name = models.CharField(max_length=50, verbose_name='Kategoriename', help_text='Name der Aufgabenkategorie')
    person_cluster = models.ManyToManyField(PersonCluster, verbose_name='Für Benutzergruppen', help_text='Benutzergruppen, für die diese Kategorie relevant ist')
    type = models.CharField(max_length=1, choices=TYPE_CHOICES, verbose_name='Zeitraum', help_text='Zeitraum, in dem Aufgaben dieser Kategorie anfallen (nur für Freiwillige)', null=True, blank=True)
    class Meta:
        verbose_name = 'Aufgabenkategorie'
        verbose_name_plural = 'Aufgabenkategorien'

    def __str__(self):
        return self.name + ' - ' + ', '.join([pc.name for pc in self.person_cluster.all()])

class Aufgabe2(OrgModel):

    name = models.CharField(max_length=50, verbose_name='Name', help_text='Name der Aufgabe')
    beschreibung = models.TextField(null=True, blank=True, verbose_name='Beschreibung', help_text='Ausführliche Beschreibung der Aufgabe')
    mitupload = models.BooleanField(default=True, verbose_name='Datei-Upload erforderlich', help_text='Ermöglicht und erfordert das Hochladen von Dateien bei dieser Aufgabe')
    requires_submission = models.BooleanField(default=True, verbose_name='Bestätigung erforderlich', help_text='Wenn aktiviert, wird die Aufgabe nach Erledigung als pending markiert und muss manuell als erledigt bestätigt werden')
    faellig_art = models.ForeignKey(AufgabenCluster, on_delete=models.SET_NULL, null=True, blank=True, verbose_name='Kategorie', help_text='Kategorie dieser Aufgabe, bestimmt den Zeitraum')
    faellig_tag = models.IntegerField(blank=True, null=True, verbose_name='Fällig am Tag', validators=[validators.MinValueValidator(1), validators.MaxValueValidator(31)], help_text='Tag im Monat, an dem die Aufgabe fällig ist (1-31)')
    faellig_monat = models.IntegerField(blank=True, null=True, verbose_name='Fällig im Monat', validators=[validators.MinValueValidator(1), validators.MaxValueValidator(12)], help_text='Monat, in dem die Aufgabe fällig ist (1-12)')
    faellig_tage_nach_start = models.IntegerField(blank=True, null=True, verbose_name='Tage nach Einsatzbeginn', validators=[validators.MinValueValidator(0)], help_text='Anzahl der Tage nach Einsatzbeginn, wann die Aufgabe fällig ist')
    faellig_tage_vor_ende = models.IntegerField(blank=True, null=True, verbose_name='Tage vor Einsatzende', validators=[validators.MinValueValidator(0)], help_text='Anzahl der Tage vor Einsatzende, wann die Aufgabe fällig ist')
    repeat_push_days = models.IntegerField(default=7, verbose_name='Wartezeit bis nächste Benachrichtigung', validators=[validators.MinValueValidator(0)], help_text='Anzahl der Tage, die gewartet wird, bevor eine neue Benachrichtigung für eine nicht erledigte Aufgabe gesendet wird')
    person_cluster = models.ManyToManyField(PersonCluster, verbose_name='Für Benutzergruppen', help_text='Benutzergruppen, für die diese Aufgabe relevant ist')
    wiederholung = models.BooleanField(default=False, verbose_name='Regelmäßige Wiederholung', help_text='Wenn aktiviert, wird die Aufgabe regelmäßig wiederholt')
    wiederholung_interval_wochen = models.IntegerField(blank=True, null=True, verbose_name='Wiederholung alle x Wochen', validators=[validators.MinValueValidator(0)], help_text='Wiederholungsintervall in Wochen')
    wiederholung_ende = models.DateField(blank=True, null=True, verbose_name='Wiederholung bis', help_text='Datum, bis zu dem die Aufgabe wiederholt wird')

    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Aufgabe'
        verbose_name_plural = 'Aufgaben'

    def __str__(self):
        return self.name
    

class AufgabeZwischenschritte2(OrgModel):
    aufgabe = models.ForeignKey(Aufgabe2, on_delete=models.CASCADE, verbose_name='Aufgabe')
    name = models.CharField(max_length=50, verbose_name='Name', help_text='Name des Zwischenschritts')
    beschreibung = models.TextField(null=True, blank=True, verbose_name='Beschreibung', help_text='Beschreibung des Zwischenschritts')

    class Meta:
        verbose_name = 'Aufgabe Zwischenschritt'
        verbose_name_plural = 'Aufgabe Zwischenschritte'

    def __str__(self):
        return self.name
    
    def save(self, *args, **kwargs):
        # First save the instance so it has an ID
        super(AufgabeZwischenschritte2, self).save(*args, **kwargs)
        
        # Now we can safely filter related objects
        user_aufgaben = UserAufgaben.objects.filter(aufgabe=self.aufgabe)
        for user_aufgabe in user_aufgaben:
            fw_aufg_zw, created = UserAufgabenZwischenschritte.objects.get_or_create(
                user_aufgabe=user_aufgabe,
                aufgabe_zwischenschritt=self,
                org=user_aufgabe.org
            )
            if created:
                fw_aufg_zw.erledigt = user_aufgabe.erledigt
                fw_aufg_zw.save()
    

class UserAufgaben(OrgModel):
    WIEDERHOLUNG_CHOICES = [
        ('T', 'Täglich'),
        ('W', 'Wöchentlich'),
        ('M', 'Monatlich'),
        ('J', 'Jährlich'),
        ('N', 'Nicht wiederholen')
    ]

    user = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name='Benutzer:in', help_text='Benutzer:in, dem diese Aufgabe zugewiesen ist', db_index=True)
    aufgabe = models.ForeignKey(Aufgabe2, on_delete=models.CASCADE, verbose_name='Aufgabe', help_text='Die zugewiesene Aufgabe', db_index=True)
    personalised_description = models.TextField(blank=True, null=True, verbose_name='Persönliche Notizen', help_text='Individuelle Anmerkungen zu dieser Aufgabe')
    erledigt = models.BooleanField(default=False, verbose_name='Abgeschlossen', help_text='Zeigt an, ob die Aufgabe abgeschlossen ist')
    pending = models.BooleanField(default=False, verbose_name='Pending', help_text='Zeigt an, dass die Aufgabe begonnen aber noch nicht abgeschlossen wurde')
    datetime = models.DateTimeField(auto_now_add=True, verbose_name='Erstellt am')
    faellig = models.DateField(blank=True, null=True, verbose_name='Fälligkeitsdatum', help_text='Datum, bis zu dem die Aufgabe erledigt sein sollte')
    last_reminder = models.DateField(blank=True, null=True, verbose_name='Letzte Erinnerung', help_text='Datum der letzten versendeten Erinnerung')
    erledigt_am = models.DateField(blank=True, null=True, verbose_name='Abgeschlossen am', help_text='Datum, an dem die Aufgabe abgeschlossen wurde')
    file = models.FileField(upload_to='uploads/', max_length=255, blank=True, null=True, verbose_name='Angehängte Datei', help_text='Datei, die für diese Aufgabe hochgeladen wurde')
    file_list = models.JSONField(blank=True, null=True, verbose_name='Angehängte Dateien', help_text='Dateien, die für diese Aufgabe hochgeladen wurden')
    file_downloaded_of = models.ManyToManyField(User, blank=True, verbose_name='Dateien heruntergeladen von', help_text='Benutzer, die die Dateien heruntergeladen haben', related_name='file_downloaded_of')
    benachrichtigung_cc = models.CharField(max_length=255, blank=True, null=True, verbose_name='E-Mail-Kopie an', help_text='Weitere E-Mail-Adressen, die Benachrichtigungen erhalten sollen (kommagetrennt)')
    history = HistoricalRecords()

    def save(self, *args, **kwargs):
        from FW.models import Freiwilliger

        if self.aufgabe.wiederholung and self.erledigt:
            # Check if there's no end date or if the current date is before the end date
            if self.aufgabe.wiederholung_ende is None or datetime.now().date() < self.aufgabe.wiederholung_ende:
                # Handle case where faellig might be None
                current_date = datetime.now().date()
                if self.faellig is None:
                    new_date = current_date
                else:
                    new_date = max(current_date, self.faellig)
                
                self.faellig = new_date + timedelta(days=self.aufgabe.wiederholung_interval_wochen * 7)
                self.erledigt = False
                self.pending = False

        if not self.faellig:

            faellig_date = datetime(datetime.now().year, self.aufgabe.faellig_monat or 1, self.aufgabe.faellig_tag or 1).date()
            while faellig_date < datetime.now().date():
                faellig_date = faellig_date.replace(year=faellig_date.year+1)

            if Freiwilliger.objects.filter(org=self.org, user=self.user, user__customuser__person_cluster__in=self.aufgabe.person_cluster.all()).exists():
                freiwilliger = Freiwilliger.objects.get(org=self.org, user=self.user, user__customuser__person_cluster__in=self.aufgabe.person_cluster.all())

                start_date = freiwilliger.start_real or freiwilliger.start_geplant
                end_date = freiwilliger.ende_real or freiwilliger.ende_geplant

                if self.aufgabe.faellig_tage_nach_start:
                    self.faellig = start_date + timedelta(days=self.aufgabe.faellig_tage_nach_start)
                elif self.aufgabe.faellig_tage_vor_ende:
                    self.faellig = end_date - timedelta(days=self.aufgabe.faellig_tage_vor_ende)
                elif self.aufgabe.faellig_monat:
                    year = start_date.year if start_date else datetime.now().year

                    try:
                        while self.aufgabe.faellig_art.type == 'V' and faellig_date > start_date:
                            faellig_date = faellig_date.replace(year=year-1)

                        while self.aufgabe.faellig_art.type == 'W' and faellig_date < start_date:
                            faellig_date = faellig_date.replace(year=year+1)
                        while self.aufgabe.faellig_art.type == 'W' and faellig_date > end_date:
                            faellig_date = faellig_date.replace(year=year-1)

                        while self.aufgabe.faellig_art.type == 'N' and faellig_date < end_date:
                                faellig_date = faellig_date.replace(year=year+1)
                    except:
                        pass

                    self.faellig = faellig_date

            elif self.aufgabe.faellig_monat:
                self.faellig = faellig_date

        
        super(UserAufgaben, self).save(*args, **kwargs)


    def send_reminder_email(self):
        from Global.send_email import send_aufgaben_email
        send_aufgaben_email(self, self.user.org)


    class Meta:
        verbose_name = 'Benutzeraufgabe'
        verbose_name_plural = 'Benutzeraufgaben'

    def __str__(self):
        return self.user.first_name + ' ' + self.user.last_name + ' - ' + self.aufgabe.name


class UserAufgabenZwischenschritte(OrgModel):
    user_aufgabe = models.ForeignKey(UserAufgaben, on_delete=models.CASCADE, verbose_name='User Aufgabe')
    aufgabe_zwischenschritt = models.ForeignKey(AufgabeZwischenschritte2, on_delete=models.CASCADE, verbose_name='Aufgabe Zwischenschritt')
    erledigt = models.BooleanField(default=False, verbose_name='Erledigt')

    class Meta:
        verbose_name = 'Freiwilliger Aufgaben Zwischenschritt'
        verbose_name_plural = 'Freiwilliger Aufgaben Zwischenschritte'

class Post2(OrgModel):
    user = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name='Benutzer', help_text='Benutzer, der den Post erstellt hat')
    title = models.CharField(max_length=50, verbose_name='Post-Titel', help_text='Titel des Posts')
    text = models.TextField(verbose_name='Text', help_text='Text des Posts')
    date = models.DateTimeField(auto_now_add=True, verbose_name='Erstellt am')
    date_updated = models.DateTimeField(auto_now=True, verbose_name='Aktualisiert am')
    has_survey = models.BooleanField(default=False, verbose_name='Umfrage', help_text='Post enthält eine Umfrage')
    person_cluster = models.ManyToManyField(PersonCluster, verbose_name='Für Benutzergruppen', help_text='Benutzergruppen, für die dieser Post relevant ist', blank=True)

    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Post'
        verbose_name_plural = 'Posts'

    def __str__(self):
        return self.title


@receiver(post_save, sender=Post2)
def send_new_post_email_task_receiver(sender, instance, created, **kwargs):
    if created:
        print(f"Sending new post email to {instance.person_cluster.all()}")
        send_new_post_email_task.apply_async(args=[instance.id], countdown=2)

class PostSurveyQuestion(OrgModel):
    post = models.OneToOneField(Post2, on_delete=models.CASCADE, related_name='survey_question', verbose_name='Post')
    question_text = models.CharField(max_length=200, verbose_name='Frage', help_text='Text der Umfragefrage')
    
    class Meta:
        verbose_name = 'Umfragefrage'
        verbose_name_plural = 'Umfragefragen'
        
    def __str__(self):
        return self.question_text


class PostSurveyAnswer(OrgModel):
    question = models.ForeignKey(PostSurveyQuestion, on_delete=models.CASCADE, related_name='survey_answers', verbose_name='Frage')
    answer_text = models.CharField(max_length=100, verbose_name='Antwort', help_text='Text der Antwortmöglichkeit')
    votes = models.ManyToManyField(User, verbose_name='Benutzer, die an der Umfrage teilgenommen haben', blank=True)
    
    class Meta:
        verbose_name = 'Umfrageantwort'
        verbose_name_plural = 'Umfrageantworten'
        
    def __str__(self):
        return self.answer_text


class Bilder2(OrgModel):
    user = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name='Benutzer')
    titel = models.CharField(max_length=50, verbose_name=_('Bildtitel'))
    beschreibung = models.TextField(blank=True, null=True, verbose_name=_('Beschreibung'))
    date_created = models.DateTimeField(auto_now_add=True, verbose_name=_('Erstellt am'))
    date_updated = models.DateTimeField(auto_now=True, verbose_name=_('Aktualisiert am'))

    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Bild'
        verbose_name_plural = 'Bilder'

    def __str__(self):
        return self.titel

    def get_comment_count(self):
        """Get total number of comments for this image."""
        from Global.models import BilderComment
        return BilderComment.objects.filter(bilder=self).count()

    def get_reaction_summary(self):
        """Get a summary of reactions grouped by emoji."""
        from django.db.models import Count
        from Global.models import BilderReaction
        possible_reactions = BilderReaction.EMOJI_CHOICES
        reactions = []
        for reaction in possible_reactions:
            reactions.append({
                'emoji': reaction[0],
                'count': BilderReaction.objects.filter(bilder=self, emoji=reaction[0]).count()
            })
        reactions.sort(key=lambda x: x['count'], reverse=True)
        return reactions
    
    def get_my_reaction(self, user):
        """Get the specified user's reaction for this image."""
        from Global.models import BilderReaction
        return BilderReaction.objects.filter(bilder=self, user=user).first()
    
    def get_all_reactions_with_users(self):
        """Get all reactions for this image grouped by emoji with user details."""
        from Global.models import BilderReaction
        reactions_by_emoji = {}
        
        for reaction in BilderReaction.objects.filter(bilder=self).select_related('user').order_by('emoji', 'date_created'):
            emoji = reaction.emoji
            if emoji not in reactions_by_emoji:
                reactions_by_emoji[emoji] = []
            
            reactions_by_emoji[emoji].append({
                'user_id': reaction.user.id,
                'user_name': reaction.user.get_full_name() or reaction.user.username,
                'date_created': reaction.date_created
            })
        
        return reactions_by_emoji

class BilderComment(OrgModel):
    bilder = models.ForeignKey(Bilder2, on_delete=models.CASCADE, related_name='comments', verbose_name='Bild')
    user = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name='Benutzer')
    comment = models.TextField(max_length=500, verbose_name=_('Kommentar'))
    date_created = models.DateTimeField(auto_now_add=True, verbose_name=_('Erstellt am'))
    date_updated = models.DateTimeField(auto_now=True, verbose_name=_('Aktualisiert am'))

    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Bildkommentar'
        verbose_name_plural = 'Bildkommentare'
        ordering = ['-date_created']

    def __str__(self):
        return f'{self.user.get_full_name()}: {self.comment[:50]}...'


class BilderReaction(OrgModel):
    EMOJI_CHOICES = [
        ('👍', _('Daumen hoch')),
        ('❤️', _('Herz')),
        ('😂', _('Lachen')),
        ('😮', _('Überrascht')),
        ('😢', _('Traurig')),
        ('😡', _('Wütend')),
    ]
    
    bilder = models.ForeignKey(Bilder2, on_delete=models.CASCADE, related_name='reactions', verbose_name='Bild')
    user = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name='Benutzer')
    emoji = models.CharField(max_length=10, choices=EMOJI_CHOICES, verbose_name=_('Emoji'))
    date_created = models.DateTimeField(auto_now_add=True, verbose_name=_('Erstellt am'))

    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Bildreaktion'
        verbose_name_plural = 'Bildreaktionen'
        unique_together = ['bilder', 'user']  # Only one reaction per user per image

    def __str__(self):
        return f'{self.user.get_full_name()}: {self.emoji} auf {self.bilder.titel}'


class BilderGallery2(OrgModel):
    image = models.ImageField(upload_to='bilder/', verbose_name='Bild')

    small_image = models.ImageField(upload_to='bilder/small/', blank=True, null=True, verbose_name='Kleines Bild')
    bilder = models.ForeignKey(Bilder2, on_delete=models.CASCADE, verbose_name='Bild')

    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Bilder Gallery'
        verbose_name_plural = 'Bilder Galleries'

    def __str__(self):
        return self.image.name
    
    def save(self, *args, **kwargs):
        if self.image and not self.small_image:
            self.small_image = calculate_small_image(self.image)
        super().save(*args, **kwargs)


@receiver(pre_delete, sender=BilderGallery2)
def delete_bilder_files(sender, instance, **kwargs):
    """Delete image files when BilderGallery instance is deleted."""
    try:
        # Delete main image file if it exists
        if instance.image:
            if os.path.isfile(instance.image.path):
                os.remove(instance.image.path)
        
        # Delete small image file if it exists
        if instance.small_image:
            if os.path.isfile(instance.small_image.path):
                os.remove(instance.small_image.path)
    except Exception as e:
        print(f"Error deleting image files: {str(e)}")



class ProfilUser2(OrgModel):
    user = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name='Benutzer')
    attribut = models.CharField(max_length=50, verbose_name='Attribut', help_text='Attribut, das gespeichert werden soll')
    value = models.TextField(verbose_name='Wert', help_text='Wert, der gespeichert werden soll')

    history = HistoricalRecords()

    class Meta:
        verbose_name = 'Profil User'
        verbose_name_plural = 'Profil User'

    def __str__(self):
        return self.user.first_name + ' ' + self.user.last_name + ' - ' + self.attribut


class Maintenance(models.Model):
    maintenance_start_time = models.DateTimeField(verbose_name='Wartung startet am')
    maintenance_end_time = models.DateTimeField(verbose_name='Wartung endet am')

    class Meta:
        verbose_name = 'Wartung'
        verbose_name_plural = 'Wartungen'


class PushSubscription(models.Model):
    """
    Model to store push notification subscription information for users.
    This enables web push notifications to be sent to browsers.
    """
    user = models.ForeignKey(User, on_delete=models.CASCADE, help_text="User who has subscribed to push notifications")
    org = models.ForeignKey(Organisation, on_delete=models.CASCADE, help_text="Organization the user belongs to")
    endpoint = models.URLField(max_length=500, help_text="Push subscription endpoint URL")
    p256dh = models.CharField(max_length=500, help_text="User's public key for encryption")
    auth = models.CharField(max_length=500, help_text="Authentication secret")
    name = models.CharField(max_length=255, blank=True, null=True, help_text="Optional name for this device/browser")
    created_at = models.DateTimeField(auto_now_add=True, help_text="When this subscription was created")
    last_used = models.DateTimeField(null=True, blank=True, help_text="When this subscription was last used successfully")
    
    class Meta:
        verbose_name = "Push-Abonnement"
        verbose_name_plural = "Push-Abonnements"
        unique_together = ('user', 'endpoint')
    
    def __str__(self):
        device_name = self.name or "Unbenanntes Gerät"
        return f"{self.user.username} - {device_name}"


class EinsatzstelleNotiz(OrgModel):
    einsatzstelle = models.ForeignKey(Einsatzstelle2, on_delete=models.CASCADE, verbose_name='Einsatzstelle')
    user = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name='Benutzer')
    notiz = models.TextField(verbose_name='Notiz', null=True, blank=True, help_text='Notiz der Einsatzstelle')
    date = models.DateTimeField(auto_now_add=True, verbose_name='Erstellt am')
    pinned = models.BooleanField(default=False, verbose_name='Angeheftet')
    
    class Meta:
        verbose_name = 'Einsatzstellen Notiz'
        verbose_name_plural = 'Einsatzstellen Notizen'

    def __str__(self):
        return f"{self.einsatzstelle.name} - {self.notiz[:10]}"
    
    
class StickyNote(OrgModel):
    user = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name='Benutzer')
    notiz = models.CharField(max_length=1000, verbose_name='Notiz', null=True, blank=True, help_text='Notiz')
    date = models.DateTimeField(auto_now_add=True, verbose_name='Erstellt am')
    pinned = models.BooleanField(default=False, verbose_name='Angeheftet')
    priority = models.IntegerField(default=0, verbose_name='Priorität', help_text='Priorität der Notiz')
    
    class Meta:
        verbose_name = 'Sticky Note'
        verbose_name_plural = 'Sticky Notes'
    
    def __str__(self):
        return f"{self.notiz[:10]}"